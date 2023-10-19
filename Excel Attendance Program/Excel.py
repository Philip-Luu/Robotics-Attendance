from openpyxl import load_workbook
from Scanning import*
from datetime import datetime

now = datetime.now()

#Constants
#First row to read
EXCEL_START_VALUE = 2
#Last row to read
EXCEL_END_VALUE = 125
#Relative path of excel file
EXCEL_RELATIVE_PATH = "Excel Attendance Program\Preseason Attendance Post-Rookie.xlsx"
#Sheet name (Should be date)
EXCEL_PAGE_NAME = str(now.month) + "-" + str(now.day)
print(EXCEL_PAGE_NAME)

#Opening File
wb = load_workbook(filename = EXCEL_RELATIVE_PATH)
sheet = wb[EXCEL_PAGE_NAME]


#Getting data from sheet 
Name_ID = []
for row in sheet.iter_rows(min_row=EXCEL_START_VALUE, max_row=EXCEL_END_VALUE):
    Name_ID.append(row[2].value)
    
scanID = True
while scanID:
    wb.save(EXCEL_RELATIVE_PATH)
    validPrompt = False
    Error = False
    while validPrompt == False:
        fullID = scanPrompt(Error, MissingPermissionSlip)
        MissingPermissionSlip = False
        if fullID == "ll":
            exit()
        try:
            index = Name_ID.index(int(fullID)) + EXCEL_START_VALUE
            if index >= 100:
                MissingPermissionSlip = True
                validPrompt = True
            else:
                validPrompt = True
        except (ValueError, TypeError):
            Error = True

    now = datetime.now()
    if sheet["D" + str(index)].value == None:
        sheet['D'+ str(index)] = now.strftime("%H:%M")
    else:
        sheet['E'+ str(index)] = now.strftime("%H:%M")
    
