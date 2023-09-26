from openpyxl import load_workbook
from Scanning import*
from datetime import datetime

now = datetime.now()

wb = load_workbook(filename = r"C:\Users\noaht\Documents\Student Attendance\Student Attendance Test.xlsx")
sheet = wb['Sheet1']


Name_ID = []
CheckedIn = []
for row in sheet.iter_rows(min_row=2, max_row=69):
    Name_ID.append(row[2].value)
    CheckedIn.append(False)
scanID = True
while scanID:
    wb.save(r"C:\Users\noaht\Documents\Student Attendance\Student Attendance Test.xlsx")
    validPrompt = False
    Error = False
    while validPrompt == False:
        fullID = scanPrompt(Error)
        if fullID == "ll":
            exit()
        try:
            index = Name_ID.index(int(fullID)) + 2
            validPrompt = True
        except ValueError:
            validPrompt = validPrompt
            Error = True
            print("Invalid Entry")

    now = datetime.now()
    print(sheet["E" + str(index)].value)
    if sheet["E" + str(index)].value == None:
        sheet['E'+ str(index)] = now.strftime("%H:%M")
        print(now.strftime("%H:%M"))
        CheckedIn[index] = True

    else:
        sheet['F'+ str(index)] = now.strftime("%H:%M")
        print(now.strftime("%H:%M"))
    
    